from datetime import datetime
from collections import Counter
from openpyxl import load_workbook

# Ielādē Excel failu
wb = load_workbook("MovieDiary.xlsx")
ws = wb.active

# Izveido jaunu lapu ar šodienas datumu
current_date = datetime.now().strftime('%Y-%m-%d')
new_ws = wb.copy_worksheet(ws)
new_ws.title = current_date

# Sagatavo mainīgos datu analīzei
film_count = 0
higher_imdb = 0
higher_user = 0
equal_scores = 0
genres = []

# Analizē katru rindu
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
    genre = row[0].value
    imdb_rating = row[2].value
    user_rating = row[3].value

    if imdb_rating is not None and user_rating is not None:
        film_count += 1
        if imdb_rating > user_rating:
            higher_imdb += 1
        elif user_rating > imdb_rating:
            higher_user += 1
        else:
            equal_scores += 1

    if genre:
        genres.append(genre)

# Populārākais žanrs
most_common_genre = Counter(genres).most_common(1)
popular_genre = most_common_genre[0][0] if most_common_genre else "N/A"

# Ieraksta rezultātus jaunajā lapā (kolonnas F un G)
summary_row = 2
new_ws[f"F{summary_row}"] = "Filmu skaits:"
new_ws[f"G{summary_row}"] = film_count
new_ws[f"F{summary_row + 1}"] = "IMDb > Tavs:"
new_ws[f"G{summary_row + 1}"] = higher_imdb
new_ws[f"F{summary_row + 2}"] = "Tavs > IMDb:"
new_ws[f"G{summary_row + 2}"] = higher_user
new_ws[f"F{summary_row + 3}"] = "Vienādi vērtējumi:"
new_ws[f"G{summary_row + 3}"] = equal_scores
new_ws[f"F{summary_row + 4}"] = "Populārākais žanrs:"
new_ws[f"G{summary_row + 4}"] = popular_genre

# Notīra iepriekšējās dienas ierakstus (saglabā galvenes)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.value = None

wb.save("MovieDiary.xlsx")
wb.close()
